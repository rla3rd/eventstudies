"""
Excel export functionality for eventstudies package.

This module provides Excel export capabilities using pandas and openpyxl.
Importing this module will add a `to_excel()` method to SingleEvent and MultipleEvents classes.
"""

import tempfile
import numpy as np
import pandas as pd
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.drawing.image import Image
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    import warnings
    warnings.warn(
        "openpyxl is required for Excel export functionality. "
        "Install it with: pip install openpyxl"
    )

from .single_event import SingleEvent
from .multiple_events import MultipleEvents


def print_table(ws, row: int, col: int, data: dict, title: str = None):
    """
    Print a table to the worksheet.
    
    Parameters
    ----------
    ws : openpyxl worksheet
        The worksheet to write to
    row : int
        Starting row (0-indexed, matching old API)
    col : int
        Starting column (0-indexed, matching old API)
    data : dict
        Dictionary with column names as keys and lists as values
    title : str, optional
        Title for the table
    
    Returns
    -------
    tuple
        (last_row, last_col) - last row and column used (0-indexed)
    """
    # Convert 0-indexed to 1-indexed for openpyxl
    row_1idx = row + 1
    col_1idx = col + 1
    
    if title:
        cell = ws.cell(row=row_1idx, column=col_1idx)
        cell.value = title
        cell.font = Font(italic=True)
        row_1idx += 1
    
    maxLen = 0
    current_col_1idx = col_1idx
    for key, values in data.items():
        # Header
        header_cell = ws.cell(row=row_1idx, column=current_col_1idx)
        header_cell.value = key
        header_cell.font = Font(bold=True)
        header_cell.border = Border(bottom=Side(style='thin'))
        
        # Data column
        for i, val in enumerate(values):
            cell = ws.cell(row=row_1idx + 1 + i, column=current_col_1idx)
            if pd.notna(val):
                if isinstance(val, (int, float)):
                    cell.value = float(val)
                else:
                    cell.value = str(val)
            else:
                cell.value = None
        
        if len(values) > maxLen:
            maxLen = len(values)
        current_col_1idx += 1
    
    # last row and column used (convert back to 0-indexed for compatibility)
    last_row = maxLen + row
    last_col = current_col_1idx - 2  # -2 because we increment then convert back (current_col_1idx - 1) - 1
    return last_row, last_col


def write_summary(self, type: str, wb, sheet_name='summary', *, chart_as_picture: bool=False):
    """
    Write summary sheet to workbook.
    
    Parameters
    ----------
    self : SingleEvent or MultipleEvents
        The event study object
    type : str
        'Single' or 'Multiple'
    wb : openpyxl Workbook
        The workbook to write to
    sheet_name : str
        Name of the sheet
    chart_as_picture : bool
        If True, insert chart as PNG image; if False, create Excel chart
    """
    ws = wb.create_sheet(title=sheet_name)
    if sheet_name == 'summary' and len(wb.worksheets) > 1:
        # Remove default sheet if we're creating summary
        wb.remove(wb.worksheets[0])
        ws = wb['summary']
    
    # Formats
    f_h1 = Font(size=15, bold=True)
    f_h2 = Font(italic=True)
    f_t_sum = Font(bold=True)
    border_right = Border(right=Side(style='thin'))
    
    # Heading (row 0, col 0 in 0-indexed = row 1, col 1 in 1-indexed)
    heading_cell = ws.cell(row=1, column=1)
    heading_cell.value = "Event study analysis"
    heading_cell.font = f_h1
    
    # Table Summary (row 0, col 0) - overwrite with "Specification"
    spec_cell = ws.cell(row=1, column=1)
    spec_cell.value = "Specification"
    spec_cell.font = f_h2
    
    # Description (row 2, col 0 and 1)
    ws.cell(row=3, column=1).value = "Description"
    ws.cell(row=3, column=1).font = f_t_sum
    ws.cell(row=3, column=1).border = border_right
    if self.description:
        ws.cell(row=3, column=2).value = self.description
    else:
        ws.cell(row=3, column=2).value = "no description"
    
    # Table of results
    if type == 'Single':
        ws.cell(row=4, column=1).value = "Event date"
        ws.cell(row=4, column=1).font = f_t_sum
        ws.cell(row=4, column=1).border = border_right
        ws.cell(row=4, column=2).value = np.datetime_as_string(self.event_date)
        
        ws.cell(row=5, column=1).value = "Event window start"
        ws.cell(row=5, column=1).font = f_t_sum
        ws.cell(row=5, column=1).border = border_right
        ws.cell(row=5, column=2).value = self.event_window[0]
        
        ws.cell(row=6, column=1).value = "Event window end"
        ws.cell(row=6, column=1).font = f_t_sum
        ws.cell(row=6, column=1).border = border_right
        ws.cell(row=6, column=2).value = self.event_window[1]
        
        ws.cell(row=7, column=1).value = "Estimation size"
        ws.cell(row=7, column=1).font = f_t_sum
        ws.cell(row=7, column=1).border = border_right
        ws.cell(row=7, column=2).value = self.est_size
        
        results = {
            "#": list(range(self.event_window[0], self.event_window[1] + 1)),
            "AR": self.AR.tolist() if hasattr(self.AR, 'tolist') else list(self.AR),
            "Variance AR": self.var_AR.tolist() if hasattr(self.var_AR, 'tolist') else list(self.var_AR),
            "CAR": self.CAR.tolist() if hasattr(self.CAR, 'tolist') else list(self.CAR),
            "Variance CAR": list(self.var_CAR) if not isinstance(self.var_CAR, list) else self.var_CAR,
            "T-stat": self.tstat.tolist() if hasattr(self.tstat, 'tolist') else list(self.tstat),
            "P-value": self.pvalue.tolist() if hasattr(self.pvalue, 'tolist') else list(self.pvalue),
        }
    elif type == 'Multiple':
        results = {
            "#": list(range(self.event_window[0], self.event_window[1] + 1)),
            "AAR": self.AAR.tolist() if hasattr(self.AAR, 'tolist') else list(self.AAR),
            "Variance AAR": self.var_AAR.tolist() if hasattr(self.var_AAR, 'tolist') else list(self.var_AAR),
            "CAAR": self.CAAR.tolist() if hasattr(self.CAAR, 'tolist') else list(self.CAAR),
            "Variance CAAR": list(self.var_CAAR) if not isinstance(self.var_CAAR, list) else self.var_CAAR,
            "T-stat": self.tstat.tolist() if hasattr(self.tstat, 'tolist') else list(self.tstat),
            "P-value": self.pvalue.tolist() if hasattr(self.pvalue, 'tolist') else list(self.pvalue),
        }
    
    # Print table starting at row 8, col 0 (0-indexed)
    last_row, last_col = print_table(ws, row=8, col=0, data=results, title="Table of results")
    
    # Display chart
    # chart_col = last_col + 2 in 0-indexed (old API)
    chart_col_0idx = last_col + 2
    chart_col_1idx = chart_col_0idx + 1  # Convert to 1-indexed for openpyxl
    
    if type == 'Single':
        ws.cell(row=3, column=chart_col_1idx).value = "Graph of CAR"
        ws.cell(row=3, column=chart_col_1idx).font = f_h2
    if type == 'Multiple':
        ws.cell(row=3, column=chart_col_1idx).value = "Graph of CAAR"
        ws.cell(row=3, column=chart_col_1idx).font = f_h2
    
    if chart_as_picture:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
            fig = self.plot(AR=True)
            fig.savefig(tmpfile.name, format="png", pad_inches=0.05, bbox_inches="tight")
            img = Image(tmpfile.name)
            # Row 4 in 0-indexed = row 5 in 1-indexed
            ws.add_image(img, f'{chr(64 + chart_col_1idx)}5')
            import matplotlib.pyplot as plt
            plt.close(fig)
    else:
        # Create Excel chart
        # CAR/CAAR line chart
        car_chart = LineChart()
        car_chart.title = "CAR" if type == 'Single' else "CAAR"
        car_chart.style = 10
        car_chart.y_axis.title = "Cumulative Abnormal Return"
        car_chart.x_axis.title = "Event Day"
        
        # Data starts at row 10 (0-indexed = row 11 in 1-indexed)
        # Categories: column 1 (the "#" column, which is col 0 in 0-indexed = col 1 in 1-indexed)
        # rows 10 to last_row (0-indexed) = rows 11 to last_row+1 (1-indexed)
        categories = Reference(ws, min_col=1, min_row=11, max_row=last_row + 1)
        car_chart.set_categories(categories)
        
        # Values: CAR/CAAR column (column 4 = index 3 in 0-indexed = column 4 in 1-indexed)
        car_col_1idx = 4  # CAR/CAAR is the 4th column (after #, AR/AAR, Variance AR/AAR)
        values = Reference(ws, min_col=car_col_1idx, min_row=10, max_row=last_row + 1)
        car_chart.add_data(values, titles_from_data=True)
        car_chart.width = 10
        car_chart.height = 6
        
        # AR column chart (only for Single)
        if type == 'Single':
            ar_chart = BarChart()
            ar_chart.type = "col"
            ar_chart.style = 10
            ar_col_1idx = 2  # AR is column 2 (after #)
            ar_values = Reference(ws, min_col=ar_col_1idx, min_row=10, max_row=last_row + 1)
            ar_chart.add_data(ar_values, titles_from_data=True)
            ar_chart.y_axis = car_chart.y_axis
            car_chart += ar_chart
        
        # Format chart
        car_chart.x_axis.position_axis = "on_tick"
        car_chart.x_axis.major_tick_mark = None
        car_chart.x_axis.line = Side(style='thin', color='000000')
        car_chart.y_axis.major_gridlines = None
        car_chart.y_axis.line = Side(style='thin', color='000000')
        car_chart.legend.position = "b"
        car_chart.plot_area.border = Border(left=Side(style='thin', color='000000'),
                                          right=Side(style='thin', color='000000'),
                                          top=Side(style='thin', color='000000'),
                                          bottom=Side(style='thin', color='000000'))
        
        # Insert chart at row 5 (0-indexed row 4 = 1-indexed row 5)
        ws.add_chart(car_chart, f'{chr(64 + chart_col_1idx)}5')


def write_Single(self, path: str, *, chart_as_picture: bool=False, event_details: bool=True):
    """
    Export SingleEvent to Excel file.
    
    Parameters
    ----------
    path : str
        Path to save the Excel file
    chart_as_picture : bool, optional
        If True, insert chart as PNG image; if False, create Excel chart (default: False)
    event_details : bool, optional
        Not used for Single events, kept for API compatibility (default: True)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    write_summary(self=self, type='Single', wb=wb, sheet_name='summary', chart_as_picture=chart_as_picture)
    wb.save(path)


def write_Multiple(self, path: str, *, chart_as_picture: bool=False, event_details: bool=True):
    """
    Export MultipleEvents to Excel file.
    
    Parameters
    ----------
    path : str
        Path to save the Excel file
    chart_as_picture : bool, optional
        If True, insert chart as PNG image; if False, create Excel chart (default: False)
    event_details : bool, optional
        If True, create individual sheets for each event (default: True)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    write_summary(self=self, type='Multiple', wb=wb, sheet_name='summary', chart_as_picture=chart_as_picture)
    if event_details and hasattr(self, 'sample'):
        for i, event in enumerate(self.sample, 1):
            write_summary(self=event, type='Single', wb=wb, sheet_name='event_'+str(i), chart_as_picture=chart_as_picture)
    wb.save(path)


# Attach methods to classes (matching old API)
SingleEvent.to_excel = write_Single
MultipleEvents.to_excel = write_Multiple

__all__ = ['SingleEvent', 'MultipleEvents']
